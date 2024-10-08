import time
import openpyxl
import requests
import os
from flask import Flask, request, render_template, jsonify, Response

app = Flask(__name__)

# Replace with your actual VirusTotal API key
API_KEY = 'a5e4da964d2817a50f41832426cf2f7bc9f21725b890362430364768798f8c9c'
url = 'https://www.virustotal.com/api/v3/domains/'
headers = {'x-apikey': API_KEY}

def get_virus_total_data_v3(domain):
    """Fetch scan data for a domain from VirusTotal API v3."""
    response = requests.get(f"{url}{domain}", headers=headers)
    if response.status_code == 200:
        return response.json()
    else:
        # Log error if there's a problem with the request
        print(f"Error fetching data for {domain}: {response.status_code}, {response.text}")
        return None

def fetch_domains(file_path):
    """Extract domains from the Excel file."""
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    domains = []
    for row in range(2, sheet.max_row + 1):  # Assuming first row is headers
        domain = sheet.cell(row=row, column=1).value
        if domain:
            # Clean the domain by removing http/https and any trailing slashes
            domain = domain.replace("https://", "").replace("http://", "").rstrip("/").strip()
            domains.append(domain)
    return domains

def process_domains_stream(file_path):
    """Process each domain and return a stream of results."""
    domains = fetch_domains(file_path)
    for index, domain in enumerate(domains):
        yield f"data: Checking domain {index + 1}/{len(domains)}: {domain}\n\n"
        result = get_virus_total_data_v3(domain)
        if result:
            malicious = result['data']['attributes']['last_analysis_stats']['malicious']
            total_scans = sum(result['data']['attributes']['last_analysis_stats'].values())
            score = f"{malicious}/{total_scans}"
            status = "malicious" if malicious > 0 else "not malicious"
            yield f"data: {domain}: {score} ({status})\n\n"
        else:
            yield f"data: {domain}: Error fetching data\n\n"

        # Wait for 20 seconds before checking the next domain
        yield f"data: Waiting 20 seconds before next scan...\n\n"
        time.sleep(20)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/check_excel', methods=['POST'])
def check_excel():
    if request.method == 'POST':
        file = request.files['file']
        if file:
            file_path = "uploaded.xlsx"
            file.save(file_path)

            # Fetch domains from the uploaded Excel file
            domains = fetch_domains(file_path)

            # Return the domains to the frontend for display
            return jsonify(domains=domains)

@app.route('/upload_and_scan', methods=['POST'])
def upload_and_scan():
    if request.method == 'POST':
        file = request.files['file']
        if file:
            file_path = "uploaded.xlsx"
            file.save(file_path)

            # Stream the domain processing to the frontend
            return Response(process_domains_stream(file_path), content_type='text/event-stream')

if __name__ == "__main__":
    # Use Render's dynamic PORT and bind to 0.0.0.0 for public access
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
